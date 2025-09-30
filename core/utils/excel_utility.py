"""
Excel Utility Module
Provides comprehensive Excel file operations for the automation framework
Supports reading, writing, and manipulating Excel files (.xlsx, .xls)
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelUtility:
    """
    Utility class for Excel file operations
    Provides methods for reading, writing, and manipulating Excel files
    """

    def __init__(self, base_path: str = None):
        """
        Initialize Excel Utility

        Args:
            base_path: Base path for Excel files. Defaults to test_data/excel/
        """
        if base_path is None:
            self.base_path = Path(__file__).parent.parent.parent.parent.parent.parent / "test_data" / "excel"
        else:
            self.base_path = Path(base_path)

        logger.info(f"ExcelUtility initialized with base path: {self.base_path}")

    # ==================== File Operations ====================

    def load_workbook(self, file_name: str, read_only: bool = False) -> openpyxl.Workbook:
        """
        Load Excel workbook

        Args:
            file_name: Name of the Excel file
            read_only: Open in read-only mode for better performance

        Returns:
            Workbook object
        """
        file_path = self.base_path / file_name

        try:
            if not file_path.exists():
                raise FileNotFoundError(f"Excel file not found: {file_path}")

            workbook = openpyxl.load_workbook(file_path, read_only=read_only, data_only=True)
            logger.info(f"Successfully loaded workbook: {file_name}")
            return workbook

        except Exception as e:
            logger.error(f"Error loading workbook {file_name}: {str(e)}")
            raise

    def create_workbook(self) -> openpyxl.Workbook:
        """
        Create new Excel workbook

        Returns:
            New Workbook object
        """
        try:
            workbook = openpyxl.Workbook()
            logger.info("Created new workbook")
            return workbook

        except Exception as e:
            logger.error(f"Error creating workbook: {str(e)}")
            raise

    def save_workbook(self, workbook: openpyxl.Workbook, file_name: str) -> bool:
        """
        Save workbook to file

        Args:
            workbook: Workbook object to save
            file_name: Output file name

        Returns:
            True if successful, False otherwise
        """
        file_path = self.base_path / file_name

        try:
            # Create directory if it doesn't exist
            file_path.parent.mkdir(parents=True, exist_ok=True)

            workbook.save(file_path)
            logger.info(f"Successfully saved workbook: {file_name}")
            return True

        except Exception as e:
            logger.error(f"Error saving workbook {file_name}: {str(e)}")
            return False

    def close_workbook(self, workbook: openpyxl.Workbook) -> None:
        """
        Close workbook

        Args:
            workbook: Workbook object to close
        """
        try:
            workbook.close()
            logger.info("Workbook closed")
        except Exception as e:
            logger.error(f"Error closing workbook: {str(e)}")

    # ==================== Sheet Operations ====================

    def get_sheet_names(self, file_name: str) -> List[str]:
        """
        Get all sheet names from workbook

        Args:
            file_name: Name of the Excel file

        Returns:
            List of sheet names
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet_names = workbook.sheetnames
            workbook.close()

            logger.info(f"Found {len(sheet_names)} sheets in {file_name}")
            return sheet_names

        except Exception as e:
            logger.error(f"Error getting sheet names: {str(e)}")
            return []

    def get_sheet(self, workbook: openpyxl.Workbook, sheet_name: str = None) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Get worksheet from workbook

        Args:
            workbook: Workbook object
            sheet_name: Name of the sheet (default: active sheet)

        Returns:
            Worksheet object
        """
        try:
            if sheet_name is None:
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]

            logger.info(f"Retrieved sheet: {sheet.title}")
            return sheet

        except Exception as e:
            logger.error(f"Error getting sheet: {str(e)}")
            raise

    def create_sheet(self, workbook: openpyxl.Workbook, sheet_name: str,
                     index: int = None) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Create new sheet in workbook

        Args:
            workbook: Workbook object
            sheet_name: Name for the new sheet
            index: Position index (default: end)

        Returns:
            New Worksheet object
        """
        try:
            if index is not None:
                sheet = workbook.create_sheet(sheet_name, index)
            else:
                sheet = workbook.create_sheet(sheet_name)

            logger.info(f"Created new sheet: {sheet_name}")
            return sheet

        except Exception as e:
            logger.error(f"Error creating sheet: {str(e)}")
            raise

    def delete_sheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> bool:
        """
        Delete sheet from workbook

        Args:
            workbook: Workbook object
            sheet_name: Name of sheet to delete

        Returns:
            True if successful, False otherwise
        """
        try:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                workbook.remove(sheet)
                logger.info(f"Deleted sheet: {sheet_name}")
                return True
            else:
                logger.warning(f"Sheet not found: {sheet_name}")
                return False

        except Exception as e:
            logger.error(f"Error deleting sheet: {str(e)}")
            return False

    def rename_sheet(self, workbook: openpyxl.Workbook, old_name: str, new_name: str) -> bool:
        """
        Rename sheet

        Args:
            workbook: Workbook object
            old_name: Current sheet name
            new_name: New sheet name

        Returns:
            True if successful, False otherwise
        """
        try:
            if old_name in workbook.sheetnames:
                sheet = workbook[old_name]
                sheet.title = new_name
                logger.info(f"Renamed sheet from {old_name} to {new_name}")
                return True
            else:
                logger.warning(f"Sheet not found: {old_name}")
                return False

        except Exception as e:
            logger.error(f"Error renaming sheet: {str(e)}")
            return False

    def copy_sheet(self, workbook: openpyxl.Workbook, sheet_name: str,
                   new_sheet_name: str) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """
        Copy sheet within workbook

        Args:
            workbook: Workbook object
            sheet_name: Name of sheet to copy
            new_sheet_name: Name for the copied sheet

        Returns:
            New Worksheet object or None
        """
        try:
            if sheet_name in workbook.sheetnames:
                source_sheet = workbook[sheet_name]
                new_sheet = workbook.copy_worksheet(source_sheet)
                new_sheet.title = new_sheet_name
                logger.info(f"Copied sheet {sheet_name} to {new_sheet_name}")
                return new_sheet
            else:
                logger.warning(f"Sheet not found: {sheet_name}")
                return None

        except Exception as e:
            logger.error(f"Error copying sheet: {str(e)}")
            return None

    # ==================== Cell Operations ====================

    def read_cell(self, file_name: str, sheet_name: str, row: int, col: int) -> Any:
        """
        Read single cell value

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number (1-based)
            col: Column number (1-based)

        Returns:
            Cell value
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            value = sheet.cell(row=row, column=col).value
            workbook.close()

            logger.info(f"Read cell ({row}, {col}): {value}")
            return value

        except Exception as e:
            logger.error(f"Error reading cell: {str(e)}")
            return None

    def write_cell(self, file_name: str, sheet_name: str, row: int, col: int, value: Any) -> bool:
        """
        Write value to single cell

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number (1-based)
            col: Column number (1-based)
            value: Value to write

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.cell(row=row, column=col).value = value
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote cell ({row}, {col}): {value}")
            return success

        except Exception as e:
            logger.error(f"Error writing cell: {str(e)}")
            return False

    def get_cell_value_by_address(self, file_name: str, sheet_name: str, cell_address: str) -> Any:
        """
        Get cell value by address (e.g., 'A1')

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_address: Cell address (e.g., 'A1', 'B5')

        Returns:
            Cell value
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            value = sheet[cell_address].value
            workbook.close()

            logger.info(f"Read cell {cell_address}: {value}")
            return value

        except Exception as e:
            logger.error(f"Error reading cell by address: {str(e)}")
            return None

    def set_cell_value_by_address(self, file_name: str, sheet_name: str,
                                  cell_address: str, value: Any) -> bool:
        """
        Set cell value by address

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_address: Cell address (e.g., 'A1')
            value: Value to write

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet[cell_address].value = value
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote cell {cell_address}: {value}")
            return success

        except Exception as e:
            logger.error(f"Error writing cell by address: {str(e)}")
            return False

    # ==================== Row Operations ====================

    def read_row(self, file_name: str, sheet_name: str, row_num: int,
                 start_col: int = 1, end_col: int = None) -> List[Any]:
        """
        Read entire row or range of columns in a row

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row_num: Row number (1-based)
            start_col: Starting column (default: 1)
            end_col: Ending column (default: last column with data)

        Returns:
            List of cell values
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            if end_col is None:
                end_col = sheet.max_column

            row_data = []
            for col in range(start_col, end_col + 1):
                value = sheet.cell(row=row_num, column=col).value
                row_data.append(value)

            workbook.close()
            logger.info(f"Read row {row_num}: {len(row_data)} cells")
            return row_data

        except Exception as e:
            logger.error(f"Error reading row: {str(e)}")
            return []

    def write_row(self, file_name: str, sheet_name: str, row_num: int,
                  data: List[Any], start_col: int = 1) -> bool:
        """
        Write data to row

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row_num: Row number (1-based)
            data: List of values to write
            start_col: Starting column (default: 1)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for idx, value in enumerate(data):
                sheet.cell(row=row_num, column=start_col + idx).value = value

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote row {row_num}: {len(data)} cells")
            return success

        except Exception as e:
            logger.error(f"Error writing row: {str(e)}")
            return False

    def insert_row(self, file_name: str, sheet_name: str, row_num: int,
                   amount: int = 1) -> bool:
        """
        Insert empty row(s)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row_num: Row number to insert at
            amount: Number of rows to insert

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.insert_rows(row_num, amount)
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Inserted {amount} row(s) at row {row_num}")
            return success

        except Exception as e:
            logger.error(f"Error inserting row: {str(e)}")
            return False

    def delete_row(self, file_name: str, sheet_name: str, row_num: int,
                   amount: int = 1) -> bool:
        """
        Delete row(s)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row_num: Row number to delete
            amount: Number of rows to delete

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.delete_rows(row_num, amount)
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Deleted {amount} row(s) starting at row {row_num}")
            return success

        except Exception as e:
            logger.error(f"Error deleting row: {str(e)}")
            return False

    def get_row_count(self, file_name: str, sheet_name: str) -> int:
        """
        Get total number of rows with data

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet

        Returns:
            Number of rows
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            row_count = sheet.max_row
            workbook.close()

            logger.info(f"Row count for {sheet_name}: {row_count}")
            return row_count

        except Exception as e:
            logger.error(f"Error getting row count: {str(e)}")
            return 0

    # ==================== Column Operations ====================

    def read_column(self, file_name: str, sheet_name: str, col_num: int,
                    start_row: int = 1, end_row: int = None) -> List[Any]:
        """
        Read entire column or range of rows in a column

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            col_num: Column number (1-based)
            start_row: Starting row (default: 1)
            end_row: Ending row (default: last row with data)

        Returns:
            List of cell values
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            if end_row is None:
                end_row = sheet.max_row

            col_data = []
            for row in range(start_row, end_row + 1):
                value = sheet.cell(row=row, column=col_num).value
                col_data.append(value)

            workbook.close()
            logger.info(f"Read column {col_num}: {len(col_data)} cells")
            return col_data

        except Exception as e:
            logger.error(f"Error reading column: {str(e)}")
            return []

    def write_column(self, file_name: str, sheet_name: str, col_num: int,
                     data: List[Any], start_row: int = 1) -> bool:
        """
        Write data to column

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            col_num: Column number (1-based)
            data: List of values to write
            start_row: Starting row (default: 1)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for idx, value in enumerate(data):
                sheet.cell(row=start_row + idx, column=col_num).value = value

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote column {col_num}: {len(data)} cells")
            return success

        except Exception as e:
            logger.error(f"Error writing column: {str(e)}")
            return False

    def insert_column(self, file_name: str, sheet_name: str, col_num: int,
                      amount: int = 1) -> bool:
        """
        Insert empty column(s)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            col_num: Column number to insert at
            amount: Number of columns to insert

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.insert_cols(col_num, amount)
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Inserted {amount} column(s) at column {col_num}")
            return success

        except Exception as e:
            logger.error(f"Error inserting column: {str(e)}")
            return False

    def delete_column(self, file_name: str, sheet_name: str, col_num: int,
                      amount: int = 1) -> bool:
        """
        Delete column(s)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            col_num: Column number to delete
            amount: Number of columns to delete

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.delete_cols(col_num, amount)
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Deleted {amount} column(s) starting at column {col_num}")
            return success

        except Exception as e:
            logger.error(f"Error deleting column: {str(e)}")
            return False

    def get_column_count(self, file_name: str, sheet_name: str) -> int:
        """
        Get total number of columns with data

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet

        Returns:
            Number of columns
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            col_count = sheet.max_column
            workbook.close()

            logger.info(f"Column count for {sheet_name}: {col_count}")
            return col_count

        except Exception as e:
            logger.error(f"Error getting column count: {str(e)}")
            return 0

    # ==================== Range Operations ====================

    def read_range(self, file_name: str, sheet_name: str, start_row: int,
                   start_col: int, end_row: int, end_col: int) -> List[List[Any]]:
        """
        Read range of cells

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row number
            start_col: Starting column number
            end_row: Ending row number
            end_col: Ending column number

        Returns:
            2D list of cell values
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            range_data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    value = sheet.cell(row=row, column=col).value
                    row_data.append(value)
                range_data.append(row_data)

            workbook.close()
            logger.info(f"Read range: {len(range_data)} rows x {len(range_data[0]) if range_data else 0} cols")
            return range_data

        except Exception as e:
            logger.error(f"Error reading range: {str(e)}")
            return []

    def write_range(self, file_name: str, sheet_name: str, start_row: int,
                    start_col: int, data: List[List[Any]]) -> bool:
        """
        Write 2D data to range

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row number
            start_col: Starting column number
            data: 2D list of values to write

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    sheet.cell(row=start_row + row_idx, column=start_col + col_idx).value = value

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote range: {len(data)} rows x {len(data[0]) if data else 0} cols")
            return success

        except Exception as e:
            logger.error(f"Error writing range: {str(e)}")
            return False

    def read_all_data(self, file_name: str, sheet_name: str,
                      skip_header: bool = False) -> List[List[Any]]:
        """
        Read all data from sheet

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            skip_header: Skip first row (header)

        Returns:
            2D list of all cell values
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            start_row = 2 if skip_header else 1
            all_data = []

            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                all_data.append(list(row))

            workbook.close()
            logger.info(f"Read all data from {sheet_name}: {len(all_data)} rows")
            return all_data

        except Exception as e:
            logger.error(f"Error reading all data: {str(e)}")
            return []

    def read_as_dict(self, file_name: str, sheet_name: str,
                     header_row: int = 1) -> List[Dict[str, Any]]:
        """
        Read sheet data as list of dictionaries (using first row as keys)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            header_row: Row number containing headers

        Returns:
            List of dictionaries
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            # Read headers
            headers = []
            for cell in sheet[header_row]:
                headers.append(cell.value)

            # Read data
            data = []
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = value
                data.append(row_dict)

            workbook.close()
            logger.info(f"Read {len(data)} rows as dictionaries from {sheet_name}")
            return data

        except Exception as e:
            logger.error(f"Error reading as dict: {str(e)}")
            return []

    # ==================== Search Operations ====================

    def find_cell(self, file_name: str, sheet_name: str, search_value: Any,
                  search_column: int = None) -> Optional[Tuple[int, int]]:
        """
        Find first cell containing value

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            search_value: Value to search for
            search_column: Specific column to search (optional)

        Returns:
            Tuple of (row, column) or None if not found
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for row in sheet.iter_rows():
                for cell in row:
                    if search_column is not None and cell.column != search_column:
                        continue
                    if cell.value == search_value:
                        result = (cell.row, cell.column)
                        workbook.close()
                        logger.info(f"Found value at: {result}")
                        return result

            workbook.close()
            logger.info(f"Value not found: {search_value}")
            return None

        except Exception as e:
            logger.error(f"Error finding cell: {str(e)}")
            return None

    def find_all_cells(self, file_name: str, sheet_name: str, search_value: Any) -> List[Tuple[int, int]]:
        """
        Find all cells containing value

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            search_value: Value to search for

        Returns:
            List of tuples (row, column)
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            results = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        results.append((cell.row, cell.column))

            workbook.close()
            logger.info(f"Found {len(results)} cells with value: {search_value}")
            return results

        except Exception as e:
            logger.error(f"Error finding cells: {str(e)}")
            return []

    def get_row_by_column_value(self, file_name: str, sheet_name: str,
                                column_name: str, value: Any,
                                header_row: int = 1) -> Optional[Dict[str, Any]]:
        """
        Get row where column contains specific value

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            column_name: Column header name
            value: Value to search for
            header_row: Row number containing headers

        Returns:
            Dictionary of row data or None
        """
        try:
            data = self.read_as_dict(file_name, sheet_name, header_row)

            for row in data:
                if row.get(column_name) == value:
                    logger.info(f"Found row where {column_name}={value}")
                    return row

            logger.info(f"No row found where {column_name}={value}")
            return None

        except Exception as e:
            logger.error(f"Error getting row by column value: {str(e)}")
            return None

    # ==================== Styling Operations ====================

    def set_cell_font(self, file_name: str, sheet_name: str, row: int, col: int,
                      font_name: str = 'Arial', font_size: int = 11,
                      bold: bool = False, italic: bool = False,
                      color: str = '000000') -> bool:
        """
        Set cell font properties

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            col: Column number
            font_name: Font name
            font_size: Font size
            bold: Bold text
            italic: Italic text
            color: Font color in hex (e.g., 'FF0000' for red)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            cell = sheet.cell(row=row, column=col)
            cell.font = Font(name=font_name, size=font_size, bold=bold,
                             italic=italic, color=color)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set font for cell ({row}, {col})")
            return success

        except Exception as e:
            logger.error(f"Error setting cell font: {str(e)}")
            return False

    def set_cell_background(self, file_name: str, sheet_name: str, row: int, col: int,
                            bg_color: str = 'FFFF00', fill_type: str = 'solid') -> bool:
        """
        Set cell background color

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            col: Column number
            bg_color: Background color in hex (e.g., 'FFFF00' for yellow)
            fill_type: Fill type ('solid', 'darkGrid', etc.)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color,
                                    fill_type=fill_type)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set background for cell ({row}, {col})")
            return success

        except Exception as e:
            logger.error(f"Error setting cell background: {str(e)}")
            return False

    def set_cell_alignment(self, file_name: str, sheet_name: str, row: int, col: int,
                           horizontal: str = 'left', vertical: str = 'top',
                           wrap_text: bool = False) -> bool:
        """
        Set cell alignment

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            col: Column number
            horizontal: Horizontal alignment ('left', 'center', 'right')
            vertical: Vertical alignment ('top', 'center', 'bottom')
            wrap_text: Enable text wrapping

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal=horizontal, vertical=vertical,
                                       wrap_text=wrap_text)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set alignment for cell ({row}, {col})")
            return success

        except Exception as e:
            logger.error(f"Error setting cell alignment: {str(e)}")
            return False

    def set_cell_border(self, file_name: str, sheet_name: str, row: int, col: int,
                        border_style: str = 'thin', border_color: str = '000000') -> bool:
        """
        Set cell border

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            col: Column number
            border_style: Border style ('thin', 'thick', 'double', etc.)
            border_color: Border color in hex

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            side = Side(style=border_style, color=border_color)
            cell = sheet.cell(row=row, column=col)
            cell.border = Border(left=side, right=side, top=side, bottom=side)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set border for cell ({row}, {col})")
            return success

        except Exception as e:
            logger.error(f"Error setting cell border: {str(e)}")
            return False

    def format_header_row(self, file_name: str, sheet_name: str, row: int = 1,
                          bg_color: str = '4472C4', font_color: str = 'FFFFFF') -> bool:
        """
        Format header row with standard styling

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Header row number
            bg_color: Background color
            font_color: Font color

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = Font(bold=True, color=font_color)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color,
                                        fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Formatted header row {row}")
            return success

        except Exception as e:
            logger.error(f"Error formatting header row: {str(e)}")
            return False

    def set_column_width(self, file_name: str, sheet_name: str, col: int, width: float) -> bool:
        """
        Set column width

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            col: Column number
            width: Column width

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = width

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set column {col} width to {width}")
            return success

        except Exception as e:
            logger.error(f"Error setting column width: {str(e)}")
            return False

    def set_row_height(self, file_name: str, sheet_name: str, row: int, height: float) -> bool:
        """
        Set row height

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            height: Row height

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            sheet.row_dimensions[row].height = height

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Set row {row} height to {height}")
            return success

        except Exception as e:
            logger.error(f"Error setting row height: {str(e)}")
            return False

    def auto_fit_columns(self, file_name: str, sheet_name: str) -> bool:
        """
        Auto-fit all column widths based on content

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                sheet.column_dimensions[column_letter].width = adjusted_width

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info("Auto-fitted all columns")
            return success

        except Exception as e:
            logger.error(f"Error auto-fitting columns: {str(e)}")
            return False

    # ==================== Formula Operations ====================

    def write_formula(self, file_name: str, sheet_name: str, row: int, col: int,
                      formula: str) -> bool:
        """
        Write formula to cell

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            row: Row number
            col: Column number
            formula: Excel formula (e.g., '=SUM(A1:A10)')

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.cell(row=row, column=col).value = formula
            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Wrote formula to cell ({row}, {col}): {formula}")
            return success

        except Exception as e:
            logger.error(f"Error writing formula: {str(e)}")
            return False

    def calculate_sum(self, file_name: str, sheet_name: str, start_row: int,
                      end_row: int, col: int, result_row: int) -> bool:
        """
        Add SUM formula to calculate column total

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row for sum
            end_row: Ending row for sum
            col: Column number
            result_row: Row to place result

        Returns:
            True if successful, False otherwise
        """
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        return self.write_formula(file_name, sheet_name, result_row, col, formula)

    def calculate_average(self, file_name: str, sheet_name: str, start_row: int,
                          end_row: int, col: int, result_row: int) -> bool:
        """
        Add AVERAGE formula to calculate column average

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            start_row: Starting row for average
            end_row: Ending row for average
            col: Column number
            result_row: Row to place result

        Returns:
            True if successful, False otherwise
        """
        col_letter = get_column_letter(col)
        formula = f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})"
        return self.write_formula(file_name, sheet_name, result_row, col, formula)

    # ==================== Data Validation ====================

    def add_dropdown_validation(self, file_name: str, sheet_name: str,
                                cell_range: str, dropdown_values: List[str]) -> bool:
        """
        Add dropdown validation to cells

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_range: Cell range (e.g., 'A1:A10')
            dropdown_values: List of dropdown options

        Returns:
            True if successful, False otherwise
        """
        try:
            from openpyxl.worksheet.datavalidation import DataValidation

            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_values)}"')
            dv.add(cell_range)
            sheet.add_data_validation(dv)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Added dropdown validation to {cell_range}")
            return success

        except Exception as e:
            logger.error(f"Error adding dropdown validation: {str(e)}")
            return False

    # ==================== Chart Operations ====================

    def create_bar_chart(self, file_name: str, sheet_name: str,
                         data_range: str, chart_title: str,
                         chart_position: str = 'E5') -> bool:
        """
        Create bar chart

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            data_range: Data range for chart (e.g., 'A1:B10')
            chart_title: Chart title
            chart_position: Position to place chart (e.g., 'E5')

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            chart = BarChart()
            chart.title = chart_title
            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            sheet.add_chart(chart, chart_position)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Created bar chart: {chart_title}")
            return success

        except Exception as e:
            logger.error(f"Error creating bar chart: {str(e)}")
            return False

    def create_line_chart(self, file_name: str, sheet_name: str,
                          data_range: str, chart_title: str,
                          chart_position: str = 'E5') -> bool:
        """
        Create line chart

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            data_range: Data range for chart
            chart_title: Chart title
            chart_position: Position to place chart

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            chart = LineChart()
            chart.title = chart_title
            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            sheet.add_chart(chart, chart_position)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Created line chart: {chart_title}")
            return success

        except Exception as e:
            logger.error(f"Error creating line chart: {str(e)}")
            return False

    def create_pie_chart(self, file_name: str, sheet_name: str,
                         data_range: str, chart_title: str,
                         chart_position: str = 'E5') -> bool:
        """
        Create pie chart

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            data_range: Data range for chart
            chart_title: Chart title
            chart_position: Position to place chart

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            chart = PieChart()
            chart.title = chart_title
            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            sheet.add_chart(chart, chart_position)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Created pie chart: {chart_title}")
            return success

        except Exception as e:
            logger.error(f"Error creating pie chart: {str(e)}")
            return False

    # ==================== Filter and Sort Operations ====================

    def enable_autofilter(self, file_name: str, sheet_name: str,
                          cell_range: str = None) -> bool:
        """
        Enable autofilter on range

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_range: Cell range (default: entire data range)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            if cell_range:
                sheet.auto_filter.ref = cell_range
            else:
                sheet.auto_filter.ref = sheet.dimensions

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info("Enabled autofilter")
            return success

        except Exception as e:
            logger.error(f"Error enabling autofilter: {str(e)}")
            return False

    def freeze_panes(self, file_name: str, sheet_name: str,
                     cell_address: str = 'A2') -> bool:
        """
        Freeze panes at specified cell

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_address: Cell where panes freeze (e.g., 'A2' freezes top row)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.freeze_panes = cell_address

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Froze panes at {cell_address}")
            return success

        except Exception as e:
            logger.error(f"Error freezing panes: {str(e)}")
            return False

    # ==================== Merge and Unmerge ====================

    def merge_cells(self, file_name: str, sheet_name: str, cell_range: str) -> bool:
        """
        Merge cells in range

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_range: Cell range to merge (e.g., 'A1:C1')

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.merge_cells(cell_range)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Merged cells: {cell_range}")
            return success

        except Exception as e:
            logger.error(f"Error merging cells: {str(e)}")
            return False

    def unmerge_cells(self, file_name: str, sheet_name: str, cell_range: str) -> bool:
        """
        Unmerge cells in range

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            cell_range: Cell range to unmerge

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)
            sheet.unmerge_cells(cell_range)

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Unmerged cells: {cell_range}")
            return success

        except Exception as e:
            logger.error(f"Error unmerging cells: {str(e)}")
            return False

    # ==================== Utility Methods ====================

    def get_column_letter_from_index(self, col_index: int) -> str:
        """
        Convert column index to letter

        Args:
            col_index: Column index (1-based)

        Returns:
            Column letter (e.g., 'A', 'B', 'AA')
        """
        return get_column_letter(col_index)

    def get_column_index_from_letter(self, col_letter: str) -> int:
        """
        Convert column letter to index

        Args:
            col_letter: Column letter (e.g., 'A', 'B', 'AA')

        Returns:
            Column index (1-based)
        """
        return column_index_from_string(col_letter)

    def clear_sheet(self, file_name: str, sheet_name: str,
                    preserve_header: bool = False) -> bool:
        """
        Clear all data from sheet

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            preserve_header: Keep first row (header)

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            start_row = 2 if preserve_header else 1

            for row in sheet.iter_rows(min_row=start_row):
                for cell in row:
                    cell.value = None

            success = self.save_workbook(workbook, file_name)
            workbook.close()

            logger.info(f"Cleared sheet: {sheet_name}")
            return success

        except Exception as e:
            logger.error(f"Error clearing sheet: {str(e)}")
            return False

    def copy_sheet_data(self, source_file: str, source_sheet: str,
                        target_file: str, target_sheet: str) -> bool:
        """
        Copy all data from one sheet to another

        Args:
            source_file: Source Excel file
            source_sheet: Source sheet name
            target_file: Target Excel file
            target_sheet: Target sheet name

        Returns:
            True if successful, False otherwise
        """
        try:
            # Read source data
            data = self.read_all_data(source_file, source_sheet)

            # Write to target
            workbook = self.load_workbook(target_file)

            if target_sheet not in workbook.sheetnames:
                self.create_sheet(workbook, target_sheet)

            sheet = self.get_sheet(workbook, target_sheet)

            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx).value = value

            success = self.save_workbook(workbook, target_file)
            workbook.close()

            logger.info(f"Copied data from {source_sheet} to {target_sheet}")
            return success

        except Exception as e:
            logger.error(f"Error copying sheet data: {str(e)}")
            return False

    def get_used_range(self, file_name: str, sheet_name: str) -> Tuple[int, int, int, int]:
        """
        Get the used range of sheet (min_row, min_col, max_row, max_col)

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet

        Returns:
            Tuple of (min_row, min_col, max_row, max_col)
        """
        try:
            workbook = self.load_workbook(file_name)
            sheet = self.get_sheet(workbook, sheet_name)

            used_range = (sheet.min_row, sheet.min_column,
                          sheet.max_row, sheet.max_column)

            workbook.close()
            logger.info(f"Used range: {used_range}")
            return used_range

        except Exception as e:
            logger.error(f"Error getting used range: {str(e)}")
            return (0, 0, 0, 0)

    def export_to_csv(self, file_name: str, sheet_name: str,
                      output_file: str) -> bool:
        """
        Export sheet data to CSV file

        Args:
            file_name: Name of the Excel file
            sheet_name: Name of the sheet
            output_file: Output CSV file path

        Returns:
            True if successful, False otherwise
        """
        try:
            import csv

            data = self.read_all_data(file_name, sheet_name)

            output_path = self.base_path.parent / "csv" / output_file
            output_path.parent.mkdir(parents=True, exist_ok=True)

            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(data)

            logger.info(f"Exported to CSV: {output_file}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            return False


# Example usage
if __name__ == "__main__":
    # Initialize Excel Utility
    excel_util = ExcelUtility()

    print("=== Excel Utility Examples ===\n")

    # Create new workbook
    wb = excel_util.create_workbook()
    sheet = wb.active
    sheet.title = "Test Data"

    # Write headers
    headers = ["ID", "Name", "Age", "Score"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    # Write data
    test_data = [
        [1, "Alice", 25, 95],
        [2, "Bob", 30, 87],
        [3, "Charlie", 28, 92],
        [4, "David", 35, 88],
        [5, "Eve", 27, 94]
    ]

    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    # Save workbook
    test_file = "test_excel.xlsx"
    excel_util.save_workbook(wb, test_file)
    print(f"Created test file: {test_file}")

    # Read data
    data = excel_util.read_all_data(test_file, "Test Data", skip_header=True)
    print(f"\nRead {len(data)} rows of data")

    # Read as dictionary
    dict_data = excel_util.read_as_dict(test_file, "Test Data")
    print(f"Read {len(dict_data)} rows as dictionaries")
    print(f"Sample row: {dict_data[0]}")

    # Format header
    excel_util.format_header_row(test_file, "Test Data")
    print("\nFormatted header row")

    # Auto-fit columns
    excel_util.auto_fit_columns(test_file, "Test Data")
    print("Auto-fitted columns")

    # Add formula
    excel_util.calculate_average(test_file, "Test Data", 2, 6, 4, 7)
    print("Added average formula")

    print("\n=== Examples completed ===")